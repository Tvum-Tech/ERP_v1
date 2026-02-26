import io
import openpyxl

from decimal import Decimal
from datetime import date
from django.utils import timezone
from django.db import transaction
from django.db.models import Sum
from django.http import HttpResponse
from django.core.exceptions import ValidationError
from django.db.models import Max
from reportlab.platypus import Image
from reportlab.lib.utils import ImageReader
import os

from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image,
)
from apps.masters.models.product import Product 
from apps.masters.models.driver import Driver
from apps.masters.models.accessory import Accessory
from collections import defaultdict

from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, Image
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from django.db.models import Sum

from apps.boq.models import BOQ, BOQItem
from apps.projects.models import Project
from apps.configurations.models import (
    LightingConfiguration,
    ConfigurationAccessory,
    ConfigurationDriver
)
class BOQPDFBuilder:
    def __init__(self, boq, is_draft=False):
        self.boq = boq
        self.is_draft = is_draft
        self.buffer = io.BytesIO()

        # Landscape A4
        self.pagesize = landscape(A4)
        self.width, self.height = self.pagesize

        self.MARGIN_X = 12 * mm
        self.MARGIN_Y = 12 * mm

        # Register Unicode font (for ₹ symbol)
        font_path = os.path.join("fonts", "DejaVuSans.ttf")
        pdfmetrics.registerFont(TTFont("DejaVu", font_path))

        self.styles = getSampleStyleSheet()

        self.style_normal = ParagraphStyle(
            'NormalSmall',
            parent=self.styles['Normal'],
            fontName='DejaVu',
            fontSize=9,
            leading=12
        )

        self.style_bold = ParagraphStyle(
            'BoldSmall',
            parent=self.styles['Normal'],
            fontName='DejaVu',
            fontSize=9,
            leading=12,
        )

    # --------------------------------------------------
    # Header + Footer
    # --------------------------------------------------
    def _header_footer(self, canvas, doc):
        canvas.saveState()

        canvas.setFont('DejaVu', 14)
        canvas.drawString(self.MARGIN_X, self.height - 30, "TVUM TECH")

        canvas.setFont('DejaVu', 10)
        canvas.drawString(self.MARGIN_X, self.height - 45, "Lighting ERP – Bill of Quantities")

        canvas.setFont('DejaVu', 9)
        canvas.drawString(self.MARGIN_X, self.height - 65, f"Project: {self.boq.project.name}")
        canvas.drawRightString(
            self.width - self.MARGIN_X,
            self.height - 65,
            f"BOQ Version: {self.boq.version}"
        )

        canvas.drawString(self.MARGIN_X, self.height - 80, f"Status: {self.boq.status}")
        canvas.drawRightString(
            self.width - self.MARGIN_X,
            self.height - 80,
            f"Date: {date.today().strftime('%d-%m-%Y')}"
        )

        canvas.setStrokeColor(colors.black)
        canvas.line(
            self.MARGIN_X,
            self.height - 90,
            self.width - self.MARGIN_X,
            self.height - 90
        )

        canvas.setFont('DejaVu', 8)
        canvas.drawCentredString(
            self.width / 2,
            15,
            "System Generated BOQ | TVUM Lighting ERP"
        )
        canvas.drawRightString(
            self.width - self.MARGIN_X,
            15,
            f"Page {doc.page}"
        )

        if self.is_draft:
            self._draw_watermark(canvas)

        canvas.restoreState()

    # --------------------------------------------------
    # Watermark
    # --------------------------------------------------
    def _draw_watermark(self, canvas):
        canvas.saveState()
        canvas.setFont('DejaVu', 50)
        canvas.setFillColor(colors.lightgrey, alpha=0.15)
        canvas.translate(self.width / 2, self.height / 2)
        canvas.rotate(45)
        canvas.drawCentredString(0, 0, "DRAFT - FOR INTERNAL USE ONLY")
        canvas.restoreState()

    # --------------------------------------------------
    # Currency
    # --------------------------------------------------
    def _format_currency(self, value):
        return f"₹ {Decimal(value):,.2f}"

    # --------------------------------------------------
    # Main Builder
    # --------------------------------------------------
    def _sr_cell(self, left_text="", right_text=""):
        inner = Table(
            [[
                Paragraph(str(left_text), self.style_normal),
                Paragraph(str(right_text), self.style_normal)
            ]],
            colWidths=[25, 25]  # Adjust if needed
        )

        inner.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))

        return inner
    def build(self):

        doc = SimpleDocTemplate(
            self.buffer,
            pagesize=self.pagesize,
            leftMargin=20 * mm,
            rightMargin=20 * mm,
            topMargin=35 * mm,
            bottomMargin=25 * mm
        )

        elements = []

        all_items = BOQItem.objects.filter(
            boq__project=self.boq.project,
            boq__version__lte=self.boq.version
        ).select_related("area", "product", "driver", "accessory")

        grouped = defaultdict(lambda: {
            "qty": Decimal(0),
            "total": Decimal(0),
            "sample": None
        })

        for item in all_items:
            key = (
                item.area_id,
                item.item_type,
                item.product_id,
                item.driver_id,
                item.accessory_id
            )
            grouped[key]["qty"] += Decimal(item.quantity or 0)
            grouped[key]["total"] += Decimal(item.final_price or 0)
            grouped[key]["sample"] = item

        area_map = defaultdict(list)
        for data in grouped.values():
            area_map[data["sample"].area].append(data)

        grand_total = Decimal(0)

        for area, items in area_map.items():

            elements.append(Paragraph(
                f"<b>Area: {area.name if area else 'General'}</b>",
                self.styles['Heading4']
            ))
            elements.append(Spacer(1, 5))

            table_data = [[
                "Sr / Sub Sr No",
                "Image",
                "Item Details",
                "Qty",
                "Unit",
                "Rate",
                "GST",
                "Total"
            ]]

            area_total = Decimal(0)
            area_subtotal = Decimal(0)
            area_gst_total = Decimal(0)
            area_final_total = Decimal(0)
            area_qty_total = Decimal(0)

            products = [d for d in items if d["sample"].item_type == "PRODUCT"]
            subitems = [d for d in items if d["sample"].item_type != "PRODUCT"]

            product_counter = 1

            for pdata in products:

                item = pdata["sample"]
                qty = pdata["qty"]
                line_total = pdata["total"]

                gst_percent = Decimal("18")

                rate = line_total / qty if qty > 0 else Decimal("0")
                subtotal = line_total
                gst_amount = subtotal * gst_percent / 100
                final_total = subtotal + gst_amount

                area_qty_total += qty
                area_subtotal += subtotal
                area_gst_total += gst_amount
                area_final_total += final_total
                area_total += final_total

                product = item.product

                details = f"<b>{product.order_code}</b><br/>{product.make} | {product.wattage}W"

                img = ""
                if product.visual_image:
                    try:
                        img = Image(product.visual_image.path, width=28, height=28)
                    except:
                        img = ""

                # Product Row
                table_data.append([
                    self._sr_cell(product_counter, ""),
                    img,
                    Paragraph(details, self.style_normal),
                    str(qty),
                    "Nos",
                    self._format_currency(rate),
                    f"{gst_percent}%",
                    self._format_currency(final_total)
                ])

                # Sub rows
                sub_counter = 1

                for sdata in subitems:
                    sitem = sdata["sample"]

                    if sitem.parent_product_id == item.product_id:

                        sqty = sdata["qty"]
                        stotal = sdata["total"]

                        srate = stotal / sqty if sqty > 0 else Decimal("0")
                        ssubtotal = stotal
                        sgst_amount = ssubtotal * gst_percent / 100
                        sfinal_total = ssubtotal + sgst_amount

                        area_qty_total += sqty
                        area_subtotal += ssubtotal
                        area_gst_total += sgst_amount
                        area_final_total += sfinal_total
                        area_total += sfinal_total

                        if sitem.item_type == "DRIVER":
                            obj = sitem.driver
                            sdesc = f"{obj.driver_make} {obj.driver_code}"
                        else:
                            obj = sitem.accessory
                            sdesc = f"{obj.accessory_name}"

                        table_data.append([
                            self._sr_cell("", f"{product_counter}.{sub_counter}"),
                            "",
                            Paragraph(sdesc, self.style_normal),
                            str(sqty),
                            "Nos",
                            self._format_currency(srate),
                            f"{gst_percent}%",
                            self._format_currency(sfinal_total)
                        ])

                        sub_counter += 1

                product_counter += 1

            # Area Total Row
            table_data.append([
                "",
                "",
                Paragraph("<b>Area Total</b>", self.style_normal),
                str(int(area_qty_total)),
                "",
                self._format_currency(area_subtotal),
                self._format_currency(area_gst_total),
                self._format_currency(area_final_total)
            ])

            usable_width = doc.width

            table = Table(
                table_data,
                colWidths=[
                    usable_width * 0.08,
                    usable_width * 0.08,
                    usable_width * 0.32,
                    usable_width * 0.07,
                    usable_width * 0.07,
                    usable_width * 0.10,
                    usable_width * 0.10,
                    usable_width * 0.18,
                ],
                repeatRows=1,
                hAlign="CENTER"
            )

            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#E8E8E8")),
                ('FONTNAME', (0, 0), (-1, -1), 'DejaVu'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#EFEFEF")),
            ]))

            elements.append(table)
            elements.append(Spacer(1, 15))

            grand_total += area_total

        elements.append(Paragraph(
            f"<b>Grand Total: {self._format_currency(grand_total)}</b>",
            ParagraphStyle(
                'Total',
                parent=self.styles['Normal'],
                alignment=TA_RIGHT,
                fontName='DejaVu',
                fontSize=13,
            )
        ))

        doc.build(
            elements,
            onFirstPage=self._header_footer,
            onLaterPages=self._header_footer
        )

        self.buffer.seek(0)
        response = HttpResponse(self.buffer, content_type="application/pdf")
        filename = f"{self.boq.project.name}_V{self.boq.version}_{self.boq.status}.pdf"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
@transaction.atomic
def generate_boq(project, user, area_id=None, subarea_id=None):

    # -----------------------------
    # 1. Load Project
    # -----------------------------
    if not isinstance(project, Project):
        project = Project.objects.get(id=project)

    # -----------------------------
    # 2. Load Active Configurations
    # -----------------------------
    active_configs = LightingConfiguration.objects.filter(
        project=project,
        is_active=True
    )

    # 🔥 Apply Area Filter (if provided)
    if area_id:
        active_configs = active_configs.filter(area_id=area_id)

    # 🔥 Apply SubArea Filter (if provided)
    if subarea_id:
        active_configs = active_configs.filter(subarea_id=subarea_id)

    active_configs = active_configs.select_related(
        "area", "subarea", "product"
    )

    if not active_configs.exists():
        raise ValidationError("No active configurations found.")

    # -----------------------------
    # 3. Detect Latest Configuration Change
    # -----------------------------
    latest_config_update = active_configs.aggregate(
        last_update=Max("updated_at")
    )["last_update"]

    latest_boq = BOQ.objects.filter(project=project).order_by("-version").first()

    if latest_boq and latest_boq.created_at >= latest_config_update:
        raise ValidationError("BOQ already generated for current configuration.")

    # -----------------------------
    # 4. Determine Next Version
    # -----------------------------
    latest_version = BOQ.objects.filter(project=project).aggregate(
        max_v=Max("version")
    )["max_v"] or 0

    next_version = latest_version + 1

    # -----------------------------
    # 5. Create BOQ Header
    # -----------------------------
    boq = BOQ.objects.create(
        project=project,
        version=next_version,
        created_by=user,
        status="DRAFT",
        source_configuration_version=next_version
    )

    # -----------------------------
    # 6. Create BOQ Items
    # -----------------------------
    seen_keys = set()

    for config in active_configs:

        area = config.area
        subarea = config.subarea
        product = config.product

        key = (
            area.id if area else None,
            subarea.id if subarea else None,
            product.pk
        )

        if key in seen_keys:
            continue

        seen_keys.add(key)

        product_price = product.base_price or 0
        product_total = product_price * config.quantity

        # PRODUCT
        BOQItem.objects.create(
            boq=boq,
            area=area,
            subarea=subarea,
            item_type="PRODUCT",
            product=product,
            parent_product=product,
            quantity=config.quantity,
            unit_price=product_price,
            markup_pct=0,
            final_price=product_total,
        )

        # DRIVERS
        drivers = ConfigurationDriver.objects.filter(configuration=config)
        for drv in drivers:
            driver_price = drv.driver.base_price or 0
            BOQItem.objects.create(
                boq=boq,
                area=area,
                subarea=subarea,
                item_type="DRIVER",
                driver=drv.driver,
                quantity=drv.quantity,
                parent_product=product,
                unit_price=driver_price,
                markup_pct=0,
                final_price=driver_price * drv.quantity,
            )

        # ACCESSORIES
        accessories = ConfigurationAccessory.objects.filter(configuration=config)
        for acc in accessories:
            acc_price = acc.accessory.base_price or 0
            BOQItem.objects.create(
                boq=boq,
                area=area,
                subarea=subarea,
                item_type="ACCESSORY",
                accessory=acc.accessory,
                parent_product=product,
                quantity=acc.quantity,
                unit_price=acc_price,
                markup_pct=0,
                final_price=acc_price * acc.quantity,
            )

    return boq
def get_boq_summary(boq):
    if not boq:
        return None

    # 🔥 cumulative items
    items = BOQItem.objects.filter(
        boq__project=boq.project,
        boq__version__lte=boq.version
    ).values("item_type").annotate(
        total_qty=Sum("quantity"),
        total_value=Sum("final_price")
    )

    summary = {}
    for item in items:
        summary[item["item_type"]] = {
            "quantity": item["total_qty"],
            "amount": float(item["total_value"] or 0)
        }

    return {
        "project_id": boq.project.id,
        "boq_id": boq.id,
        "version": boq.version,
        "status": boq.status,
        "summary": summary,
        "created_at": boq.created_at,
        "source_configuration_version": boq.source_configuration_version
    }


def get_project_boq_summary(project):
    boq = BOQ.objects.filter(project=project).order_by("-version").first()
    return get_boq_summary(boq)

def approve_boq(boq):
    if boq.status != "DRAFT":
        raise ValidationError("Already finalized")
    boq.status = "FINAL"
    boq.locked_at = timezone.now()
    boq.save()
    return boq

def apply_margin_to_boq(boq, markup_pct):
    if boq.status != "DRAFT":
        raise ValidationError("Cannot modify FINAL BOQ")
    markup_pct = Decimal(markup_pct)
    for item in boq.items.all():
        item.markup_pct = markup_pct
        item.final_price = Decimal(item.unit_price) * Decimal(item.quantity) * (Decimal(1) + markup_pct / Decimal(100))
        item.save(update_fields=["markup_pct", "final_price"])
    return boq

class BOQExcelBuilder:
    def __init__(self, boq):
        self.boq = boq
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "BOQ"
        self.row = 1

    def money(self, value): return float(round(Decimal(value), 2))
    def bold(self): return openpyxl.styles.Font(bold=True)
    def center(self): return openpyxl.styles.Alignment(horizontal="center")
    def right(self): return openpyxl.styles.Alignment(horizontal="right")

    def write(self, col, value, bold=False, align=None, currency=False):
        cell = self.ws.cell(row=self.row, column=col, value=value)
        if bold: cell.font = self.bold()
        if align: cell.alignment = align
        if currency: cell.number_format = '₹#,##0.00'
        return cell

    def build(self):
        if self.boq.status != "FINAL": raise ValidationError("Excel export is allowed only for FINAL BOQ")
        self.write(1, "TVUM TECH", bold=True)
        self.row += 1
        self.write(1, "Lighting ERP – Bill of Quantities", bold=True)
        self.row += 2
        self.write(1, "Project:", bold=True); self.write(2, self.boq.project.name)
        self.write(5, "Version:", bold=True); self.write(6, self.boq.version)
        self.row += 1
        self.write(1, "Status:", bold=True); self.write(2, self.boq.status)
        self.write(5, "Date:", bold=True); self.write(6, date.today().strftime("%d-%m-%Y"))
        self.row += 2
        grand_total = Decimal(0)
        areas = self.boq.items.select_related("area").order_by("area__name").values_list("area__id", "area__name").distinct()
        for area_id, area_name in areas:
            self.write(1, f"Area: {area_name}", bold=True); self.row += 1
            headers =[
    "Image",
    "Type",
    "Code",
    "Description",
    "Qty",
    "Unit",
    "Rate",
    "GST",
    "Total"
]
            for col, h in enumerate(headers, start=1): self.write(col, h, bold=True, align=self.center())
            self.row += 1
            area_total = Decimal(0)
            for item in self.boq.items.filter(area_id=area_id):
                qty = Decimal(item.quantity)
                unit_price = Decimal(item.unit_price or 0)
                margin = Decimal(item.markup_pct or 0)
                line_total = Decimal(item.final_price or 0)
                area_total += line_total
                grand_total += line_total
                desc = "-"
                if item.item_type == "PRODUCT" and item.product: desc = item.product.make
                elif item.item_type == "DRIVER" and item.driver: desc = item.driver.driver_type
                elif item.item_type == "ACCESSORY" and item.accessory: desc = item.accessory.accessory_type
                self.write(1, item.item_type)
                self.write(2, item.product.order_code if item.product else "-")
                self.write(3, desc)
                self.write(4, int(qty), align=self.center())
                self.write(5, self.money(unit_price), align=self.right(), currency=True)
                self.write(6, float(margin), align=self.center())
                self.write(7, self.money(line_total), align=self.right(), currency=True)
                self.row += 1
            self.write(6, "Area Total", bold=True, align=self.right())
            self.write(7, self.money(area_total), bold=True, currency=True)
            self.row += 2
        self.write(6, "Grand Total", bold=True, align=self.right())
        self.write(7, self.money(grand_total), bold=True, currency=True)
        for col in range(1, 8): self.ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="BOQ_{self.boq.project.name}_V{self.boq.version}.xlsx"'
        self.wb.save(response)
        return response